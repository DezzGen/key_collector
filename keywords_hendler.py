import openpyxl
import datetime
import os

from colorama import init
from colorama import Fore, Back, Style

from selenium import webdriver
from selenium.webdriver.chrome.options import Options

from bs4 import BeautifulSoup

from tqdm import tqdm

init()

def mixing_keywords(arr):
	temp_dict = {}
	for arr_1 in arr:
		if arr_1[0] in temp_dict:
			temp_dict[arr_1[0]].append([arr_1[1], arr_1[2]])
		else:
			temp_dict[arr_1[0]] = []
			temp_dict[arr_1[0]].append([arr_1[1], arr_1[2]])
	return temp_dict


def keycollectorseo(filename):
	#открываем файл
	workbook_read = openpyxl.load_workbook(filename, read_only=True)

	# получить доступ к отдельному листу
	first_sheet = workbook_read.worksheets[0]

	not_relevent_word_list = []
	data_ya = []
	data_go = []
	general_dict = {}

	i = 0

	# что бы перебрать строки на листе, используем
	for row in first_sheet.rows:
		# пропускаем верхнюю строку
		if i == 0:
			i = i + 1
			continue

		if str(row[75].value) == '-1' and str(row[76].value) == '-' and str(row[77].value) == '-1' and str(row[78].value) == '-':
			not_relevent_word_list.append(row[1].value)
		else:
			if str(row[76].value) != "-":
				data_ya.append([row[76].value, row[1].value, row[75].value]) # по яндекс
			if str(row[78].value) != "-":
				data_go.append([row[78].value, row[1].value, row[77].value]) # по гугл

	    # print(row[1].value)  # ключ
	    # print(row[75].value) # позиция Яндекс
	    # print(row[76].value) # страница в поиске
	    # print(row[77].value) # позиция Google -1
	    # print(row[78].value) # страница в поиске -


	data_ya = mixing_keywords(data_ya)
	data_go = mixing_keywords(data_go)

	for key_ya, value_ya in data_ya.items():
		for element_ya in value_ya:
			if key_ya in general_dict:
				general_dict[key_ya].append([element_ya[0], element_ya[1], 'None'])
			else:
				general_dict[key_ya] = []
				general_dict[key_ya].append([element_ya[0], element_ya[1], 'None'])


	for key_go, value_go in data_go.items():
		for element_go in value_go:
			
			if key_go in general_dict:
				# проходить столько раз сколько есть совпадений в гугл списке
				iii = 0
				check = element_go[0]
				for temp_el in general_dict[key_go]:
					if str(element_go[0]) == str(temp_el[0]):
						general_dict[key_go][iii][2] = element_go[1]
						check = 1;
					iii = iii + 1
				if check != 1:
					general_dict[key_go].append([element_go[0], 'None', element_go[1]])
			else:
				general_dict[key_go] = []
				general_dict[key_go].append([element_go[0], 'None', element_go[1]])

	# создаю новую книгу
	workbook = openpyxl.Workbook()

	# выбираем активный лист и меняем ему название
	ws_1 = workbook.active
	ws_1.title = "Релевантные страницы"






	options = Options()
	# Запускаем драйвер без графической оболочки браузера
	options.headless = True
	# Убираем логирование в консоль
	options.add_argument('--log-level=3')
	# Инициализируем драйвер хром
	driver = webdriver.Chrome(chrome_options=options, executable_path='drivers/chromedriver.exe')
	# Затираем первую строку
	# os.system("cls")

	# Заполняем основную таблицу
	ws_1['A'+str(1)] = 'Страница'
	ws_1['B'+str(1)] = 'Запрос'
	ws_1['C'+str(1)] = 'Позиция в Гугл'
	ws_1['D'+str(1)] = 'Позиция в Яндекс'
	ws_1['E'+str(1)] = 'TITLE'
	ws_1['F'+str(1)] = 'DESCRIPTION'


	progress_line = len(general_dict) + 2
	pbar = tqdm(total=progress_line)
	
	iii = 2
	i_pbar = 1
	for key, value in general_dict.items():
		# Получаем title и description
		driver.get(key)
		soup = BeautifulSoup (driver.page_source, features="html.parser")
		title = soup.title.string
		description = soup.find('meta', {'name':'description'})
		if description != None:
			description = description.get('content')
		pbar.update(i_pbar)
		i_pbar = i_pbar + 1
		# Заполняем таблицу данными
		ws_1['A'+str(iii)] = key
		ws_1['E'+str(iii)] = title
		ws_1['F'+str(iii)] = description

		for site_page in value:
			ws_1['B'+str(iii)] = site_page[0]
			ws_1['C'+str(iii)] = site_page[1]
			ws_1['D'+str(iii)] = site_page[2]
			iii = iii + 1

	driver.close()

	# создание нового листа
	ws_2 = workbook.create_sheet('Не релевантные запросы', 1)

	ii = 1
	for element in not_relevent_word_list:
		ws_2['A'+str(ii)] = element
		ii = ii + 1
	
	pbar.update(progress_line - 1)

	newfilename = datetime.datetime.today().strftime("%Y-%m-%d-%H-%M-%S")
	# сохраняем файл
	workbook.save(newfilename+'.xlsx')
	
	pbar.update(progress_line)
	pbar.close()



def choose_file():

	def choose_one_file():
		while True:
			input_data = input("Какой файл необходимо обработать: ")
			if not input_data.isnumeric():
				print("Вы ввели не число. Попробуйте снова")
			elif int(input_data) in list_files:
				return input_data
				break
			elif int(input_data) == 0:
				print(Fore.WHITE)
				print(Back.BLACK)
				os.abort()
			else:
				print("Ваше число в неправильном диапазоне. Попробуйте снова")


	print(Fore.BLACK)
	print(Back.YELLOW)

	list_files = {}

	i = 1

	print(' 0 - Выход ')
	for root, dirs, files in os.walk("."):
		for filename in files:
			if root == '.':
				list_files[i] = filename
				print(' ' +str(i) + ' - ' +filename)
			i += 1

	print(Fore.WHITE)
	print(Back.GREEN)

	key_choose_file = choose_one_file()
	return list_files[int(key_choose_file)]




def menu():
	print(Fore.BLACK)
	print(Back.YELLOW)

	print(' 1 - Обработать KeyCollector SEO позиции ')
	print(' 2 - Удалить ключи с вхождение ')
	print(' 3 - Собрать META данные страниц ')
	print(' 0 - Выход ')

	print(Fore.WHITE)
	print(Back.GREEN)

	q = input(' Выбери действие ? ')

	if q == '1':
		filename = choose_file()
		keycollectorseo(filename)
	elif q == '2':
		print(Fore.BLACK)
		print(Back.YELLOW)
		
		print(' 1 - Удалить слова из файла')
		print(' 2 - Задать слово')


		while True:
			print(Fore.WHITE)
			print(Back.GREEN)
			input_data = input("Какой вариант выбрать: ")
			if not input_data.isnumeric():
				print("Вы ввели не число. Попробуйте снова")
			elif int(input_data) != 1 or int(input_data) != 2:
				choose = int(input_data)
				break
			else:
				print("Ваше число в неправильном диапазоне. Попробуйте снова")
		
		if choose == 1 :
			filename = choose_file()
			print(filename)
		elif choose == 2:
			print()
			print()
			delete_word = input('Введи слово : ')
	
	elif q == '3':
		print('Выбран третий вариант')
	elif q == 's':
		pass
	elif q == '0':
		pass
	else:
		menu()



menu()